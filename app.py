import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io

st.set_page_config(page_title="Smart Scheduler SMPN 2 Banguntapan", layout="wide")

list_hari = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat']
tingkatan = ['7', '8', '9']
abjad = ['A', 'B', 'C', 'D', 'E']
semua_kelas = [f"{t}{a}" for t in tingkatan for a in abjad]

def ambil_tiga_digit(nama_kode):
    if pd.isna(nama_kode) or not nama_kode:
        return ""
    nama_str = str(nama_kode).strip()
    return nama_str[:3].upper()

def get_default_data():
    # Data diperbarui total berdasarkan gambar yang diunggah pengguna
    raw_data = [
        {"No": 1, "Kode Guru": "Purwanto", "Nama Guru": "Purwanto, S.Pd.", "Mata Pelajaran": "IPA", "Kode Mapel": "IPA", "JP/Minggu": 5, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E"], "Hari Libur/MGMP": ["Jumat"]},
        {"No": 2, "Kode Guru": "Nurkhasanah", "Nama Guru": "Nurkhasanah, S.Ag.", "Mata Pelajaran": "P.A. ISLAM", "Kode Mapel": "AGM", "JP/Minggu": 3, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 3, "Kode Guru": "Heni P", "Nama Guru": "Heni Purwaningsih, S.Pd.", "Mata Pelajaran": "B. INGGRIS", "Kode Mapel": "ING", "JP/Minggu": 5, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Selasa"]},
        {"No": 4, "Kode Guru": "Sri Purwanti", "Nama Guru": "Sri Purwanti, S.Pd.", "Mata Pelajaran": "MATEMATIKA", "Kode Mapel": "MAT", "JP/Minggu": 5, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Kamis"]},
        {"No": 5, "Kode Guru": "Rizka D", "Nama Guru": "Rizka Diestriana, S.Pd.", "Mata Pelajaran": "B. INGGRIS", "Kode Mapel": "ING", "JP/Minggu": 5, "Mengajar Kelas": ["8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Selasa"]},
        {"No": 6, "Kode Guru": "Anik M", "Nama Guru": "Anik Mulyani, S.Ag.", "Mata Pelajaran": "P.A. ISLAM", "Kode Mapel": "AGM", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 7, "Kode Guru": "Mersi", "Nama Guru": "Mersi, S.T.", "Mata Pelajaran": "INFORMATIKA", "Kode Mapel": "INF", "JP/Minggu": 3, "Mengajar Kelas": ["8A", "8B", "8C", "9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Kamis"]},
        {"No": 8, "Kode Guru": "Fitri L", "Nama Guru": "Fitri Lestari, S.S.", "Mata Pelajaran": "B. JAWA", "Kode Mapel": "JW", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E", "8A", "8B", "9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Selasa"]},
        {"No": 9, "Kode Guru": "Yoma S", "Nama Guru": "Yoma Septiantika, S.Pd.", "Mata Pelajaran": "SENI BUDAYA", "Kode Mapel": "SB", "JP/Minggu": 3, "Mengajar Kelas": ["8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Kamis"]},
        {"No": 10, "Kode Guru": "Maftuhah", "Nama Guru": "Maftuhah Rahayu, S.Pd.", "Mata Pelajaran": "B. INDONESIA", "Kode Mapel": "IND", "JP/Minggu": 5, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Selasa"]},
        {"No": 11, "Kode Guru": "Nandar P", "Nama Guru": "Nandar Pamungkas Sari, S.Pd.", "Mata Pelajaran": "BK", "Kode Mapel": "BK", "JP/Minggu": 0, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 12, "Kode Guru": "Darpito", "Nama Guru": "Darpito Nugroho, S.Pd.", "Mata Pelajaran": "BK", "Kode Mapel": "BK", "JP/Minggu": 0, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 13, "Kode Guru": "Asti A", "Nama Guru": "Asti Am Rini, S.Pd.", "Mata Pelajaran": "B. INDONESIA", "Kode Mapel": "IND", "JP/Minggu": 5, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E"], "Hari Libur/MGMP": ["Selasa"]},
        {"No": 14, "Kode Guru": "Cholid D", "Nama Guru": "Cholid Dalyanto, S.Pd.Kor.", "Mata Pelajaran": "PJOK", "Kode Mapel": "PJOK", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "7B", "7C", "8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 15, "Kode Guru": "Feni D", "Nama Guru": "Feni Dwimartanti, S.Pd.", "Mata Pelajaran": "PEND. PANCASILA", "Kode Mapel": "PP", "JP/Minggu": 4, "Mengajar Kelas": ["7C", "7D", "7E", "9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Senin"]},
        {"No": 16, "Kode Guru": "Ali S", "Nama Guru": "Ali Sudrajat, S.Pd.", "Mata Pelajaran": "BK", "Kode Mapel": "BK", "JP/Minggu": 0, "Mengajar Kelas": ["8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 17, "Kode Guru": "Rahmat M", "Nama Guru": "Rahmat Mas Said, S.Pd.", "Mata Pelajaran": "SENI BUDAYA", "Kode Mapel": "SB", "JP/Minggu": 3, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Kamis"]},
        {"No": 18, "Kode Guru": "Ami R", "Nama Guru": "Ami Royati, S.Pd.", "Mata Pelajaran": "MATEMATIKA", "Kode Mapel": "MAT", "JP/Minggu": 5, "Mengajar Kelas": ["8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Kamis"]},
        {"No": 19, "Kode Guru": "Umi K", "Nama Guru": "Umi Kulstum, S.Pd.", "Mata Pelajaran": "IPA", "Kode Mapel": "IPA", "JP/Minggu": 5, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Jumat"]},
        {"No": 20, "Kode Guru": "Eny W", "Nama Guru": "Eny Widiyanti, S.Pd.", "Mata Pelajaran": "MATEMATIKA", "Kode Mapel": "MAT", "JP/Minggu": 5, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E"], "Hari Libur/MGMP": ["Kamis"]},
        {"No": 21, "Kode Guru": "Budi P", "Nama Guru": "Budi Prasetya, S.Pd.", "Mata Pelajaran": "IPS", "Kode Mapel": "IPS", "JP/Minggu": 4, "Mengajar Kelas": ["9A", "9B", "9C", "9D", "9E", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 22, "Kode Guru": "Iswarasan", "Nama Guru": "Iswarasanti, S.Ag.", "Mata Pelajaran": "P.A. Hindu", "Kode Mapel": "AGM", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "8C", "9A"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 23, "Kode Guru": "Sabti H", "Nama Guru": "Sabti Herma Nugraheni, S.Pd.", "Mata Pelajaran": "P.A. Katolik", "Kode Mapel": "AGM", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "8A", "9A"], "Hari Libur/MGMP": ["Jumat"]},
        {"No": 24, "Kode Guru": "Intan W", "Nama Guru": "Intan Widiastuti, S.Th.", "Mata Pelajaran": "P.A. Kristen", "Kode Mapel": "AGM", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "8A", "9A"], "Hari Libur/MGMP": ["Jumat"]},
        {"No": 25, "Kode Guru": "Christina I", "Nama Guru": "Christina Dwi Ayu Wijaya, S.Pd.", "Mata Pelajaran": "PEND. PANCASILA", "Kode Mapel": "PP", "JP/Minggu": 4, "Mengajar Kelas": ["7A", "7B", "8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Senin"]},
        {"No": 26, "Kode Guru": "Thiara M", "Nama Guru": "Thiara Maharani, S.Pd.", "Mata Pelajaran": "PRAKARYA", "Kode Mapel": "PRAK", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E"], "Hari Libur/MGMP": ["Senin"]},
        {"No": 27, "Kode Guru": "Thiara M", "Nama Guru": "Thiara Maharani, S.Pd.", "Mata Pelajaran": "B.JAWA", "Kode Mapel": "JW", "JP/Minggu": 3, "Mengajar Kelas": ["8C", "8D", "8E"], "Hari Libur/MGMP": ["Senin"]},
        {"No": 28, "Kode Guru": "Luthfan Q", "Nama Guru": "Luthfan Qaedi Wicaksono, S.Pd.", "Mata Pelajaran": "PJOK", "Kode Mapel": "PJOK", "JP/Minggu": 3, "Mengajar Kelas": ["7D", "7E", "9A", "9B", "9C", "9D", "9E"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 29, "Kode Guru": "Anggiyani", "Nama Guru": "Anggiyani Fabilah Parwati, S.Pd.", "Mata Pelajaran": "INFORMATIKA", "Kode Mapel": "INF", "JP/Minggu": 3, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E", "8D", "8E"], "Hari Libur/MGMP": ["Kamis"]},
        {"No": 30, "Kode Guru": "Wesda A", "Nama Guru": "Wesda Ayu Rahmadani", "Mata Pelajaran": "B. INGGRIS", "Kode Mapel": "ING", "JP/Minggu": 5, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E"], "Hari Libur/MGMP": ["Selasa"]},
        {"No": 31, "Kode Guru": "Anisa S", "Nama Guru": "Anisa Safera Proborini, S.Pd.", "Mata Pelajaran": "IPA", "Kode Mapel": "IPA", "JP/Minggu": 5, "Mengajar Kelas": ["8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Jumat"]},
        {"No": 32, "Kode Guru": "Krisma D", "Nama Guru": "Krisma Dewi, S.Pd.", "Mata Pelajaran": "B. INDONESIA", "Kode Mapel": "IND", "JP/Minggu": 5, "Mengajar Kelas": ["8A", "8B", "8C", "8D", "8E"], "Hari Libur/MGMP": ["Selasa"]},
        {"No": 33, "Kode Guru": "Anggi S", "Nama Guru": "Anggi Supriyadi, S.Hum", "Mata Pelajaran": "P.A. Islam", "Kode Mapel": "AGM", "JP/Minggu": 3, "Mengajar Kelas": ["8A", "8B"], "Hari Libur/MGMP": ["Rabu"]},
        {"No": 34, "Kode Guru": "Rizal R", "Nama Guru": "Rizal Rahmanto, S.Pd.", "Mata Pelajaran": "IPS", "Kode Mapel": "IPS", "JP/Minggu": 4, "Mengajar Kelas": ["7A", "7B", "7C", "7D", "7E", "8A", "8B"], "Hari Libur/MGMP": ["Rabu"]}
    ]
    df_init = pd.DataFrame(raw_data)
    df_init['Banyak Kelas'] = df_init['Mengajar Kelas'].apply(len)
    df_init['Total JP Sepekan'] = df_init['JP/Minggu'] * df_init['Banyak Kelas']
    return df_init[["No", "Kode Guru", "Nama Guru", "Mata Pelajaran", "Kode Mapel", "JP/Minggu", "Mengajar Kelas", "Hari Libur/MGMP"]]

if 'data_beban_guru' not in st.session_state:
    st.session_state.data_beban_guru = get_default_data()

if 'jadwal_terplot' not in st.session_state:
    st.session_state.jadwal_terplot = []

def export_template_user(df_guru):
    # Membuat template download yang sama persis formatnya dengan gambar user
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Data Guru"
    headers = ["No", "Kode Guru", "Nama Guru", "Mata Pelajaran", "Kode Mapel", "JP/Minggu", "Mengajar Kelas", "Hari Libur/MGMP"]
    
    for col_idx, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=text)
        cell.font = Font(bold=True)
        
    for r_idx, row in df_guru.iterrows():
        ws.cell(row=r_idx+2, column=1, value=row["No"])
        ws.cell(row=r_idx+2, column=2, value=row["Kode Guru"])
        ws.cell(row=r_idx+2, column=3, value=row["Nama Guru"])
        ws.cell(row=r_idx+2, column=4, value=row["Mata Pelajaran"])
        ws.cell(row=r_idx+2, column=5, value=row["Kode Mapel"])
        ws.cell(row=r_idx+2, column=6, value=row["JP/Minggu"])
        
        # Kelas diubah menjadi teks string dipisah koma (contoh: 7A,7B,7C)
        kelas_str = ",".join(row["Mengajar Kelas"])
        ws.cell(row=r_idx+2, column=7, value=kelas_str)
        
        libur_str = ",".join(row["Hari Libur/MGMP"])
        ws.cell(row=r_idx+2, column=8, value=libur_str)
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

def export_excel_laporan(plotting_data, kelas_list, hari_list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Jadwal"
    ws.views.sheetView[0].showGridLines = True
    
    fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_subhead = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")
    fill_day = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    thin_border = Border(left=Side(style='thin', color='D1D5DB'), right=Side(style='thin', color='D1D5DB'), top=Side(style='thin', color='D1D5DB'), bottom=Side(style='thin', color='D1D5DB'))
    
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(kelas_list) + 2)
    ws["A1"] = "JADWAL TIMETABLE LENGKAP - SMPN 2 BANGUNTAPAN"
    ws["A1"].font = Font(name="Arial", size=13, bold=True, color="1B365D")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    ws.cell(row=3, column=1, value="HARI").fill = fill_header
    ws.cell(row=3, column=2, value="JAM").fill = fill_header
    ws.cell(row=3, column=1).font = Font(color="FFFFFF", bold=True)
    ws.cell(row=3, column=2).font = Font(color="FFFFFF", bold=True)
    
    for c_idx, kelas in enumerate(kelas_list, 3):
        cell = ws.cell(row=3, column=c_idx, value=f"KLS {kelas}")
        cell.font = Font(color="FFFFFF", bold=True, size=10)
        cell.fill = fill_subhead
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
        
    current_row = 4
    for hari in hari_list:
        max_jam = 6 if hari == 'Jumat' else 9
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + max_jam - 1, end_column=1)
        day_cell = ws.cell(row=current_row, column=1, value=hari.upper())
        day_cell.fill = fill_day
        day_cell.font = Font(bold=True)
        day_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for jam in range(1, max_jam + 1):
            r_num = current_row + jam - 1
            ws.cell(row=r_num, column=2, value=jam).border = thin_border
            ws.cell(row=r_num, column=1).border = thin_border
            
            for c_idx, kelas in enumerate(kelas_list, 3):
                cell_kbm = ws.cell(row=r_num, column=c_idx)
                cell_kbm.border = thin_border
                cell_kbm.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                if hari == 'Senin' and jam == 1:
                    cell_kbm.value = "UPACARA"
                    continue
                    
                matches = [d for d in plotting_data if d['hari'] == hari and d['jam'] == jam and d['kelas'] == kelas]
                if matches:
                    if len(matches) > 1 and "AGM" in matches[0]['kode_mapel']:
                        cell_kbm.value = "AGM" + chr(10) + "(PAR)"
                    else:
                        g_3d = ambil_tiga_digit(matches[0]['kode_guru'])
                        cell_kbm.value = f"{matches[0]['kode_mapel']}" + chr(10) + f"({g_3d})"
                        
        current_row += max_jam
        
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

st.title("🏫 AI Smart Timetable - Versi Optimasi Khusus")
st.subheader("SMPN 2 Banguntapan — Penempatan Pagi Mapel Esensial & Audit Validasi Kurikulum")

with st.sidebar:
    st.markdown("### 📥 Import Data Guru via Excel")
    uploaded_file = st.file_uploader("Unggah Master Data Kerja (.xlsx)", type=["xlsx"])
    
    # Tombol template unduh menggunakan format teks bersih dipisah koma sesuai keinginan user
    template_data = export_template_user(get_default_data())
    st.download_button(
        label="📄 Download Template Format Excel",
        data=template_data,
        file_name="Template_Data_Guru.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if uploaded_file is not None:
        try:
            df_upload = pd.read_excel(uploaded_file)
            if 'Mengajar Kelas' in df_upload.columns:
                df_upload['Mengajar Kelas'] = df_upload['Mengajar Kelas'].apply(
                    lambda x: [k.strip() for k in str(x).split(',')] if ',' in str(x) else [str(x).strip()]
                )
            if 'Hari Libur/MGMP' in df_upload.columns:
                df_upload['Hari Libur/MGMP'] = df_upload['Hari Libur/MGMP'].apply(
                    lambda x: [h.strip() for h in str(x).split(',')] if pd.notna(x) and ',' in str(x) else ([] if pd.isna(x) or str(x).strip()=="" else [str(x).strip()])
                )
            st.session_state.data_beban_guru = df_upload
            st.sidebar.success("✔️ File Excel Berhasil Di-import!")
        except Exception as e:
            st.sidebar.error(f"Gagal membaca file: {e}")

tab1, tab2, tab3 = st.tabs(["📋 Data Beban Mengajar Guru", "🗓️ Hasil Peta Jadwal KBM", "🔍 Laporan Audit Keselarasan Kelas"])

with tab1:
    st.info("💡 Data di bawah ini otomatis terisi berdasarkan gambar tabel baru Anda. Anda bisa mengedit barisnya langsung.")
    edited_df = st.data_editor(st.session_state.data_beban_guru, use_container_width=True)
    st.session_state.data_beban_guru = edited_df

    if st.button("⚡ Jalankan Engine AI Schedule Optimizer", type="primary"):
        plot_res = []
        
        # 1. PETA PARALEL AGAMA
        df_agama = edited_df[edited_df['Mata Pelajaran'].str.contains("P.A.", case=False, na=False)]
        hari_agama_map = {"7": ("Rabu", 4), "8": ("Kamis", 4), "9": ("Selasa", 4)}
        
        for kelas in semua_kelas:
            t = kelas[0]
            h_opt, j_opt = hari_agama_map[t]
            guru_agama_rombel = df_agama[df_agama['Mengajar Kelas'].apply(lambda x: kelas in x if isinstance(x, list) else False)]
            for _, g_row in guru_agama_rombel.iterrows():
                for step in range(3):
                    plot_res.append({
                        "kode_guru": g_row['Kode Guru'], "kode_mapel": "AGM",
                        "kelas": kelas, "hari": h_opt, "jam": j_opt + step
                    })
                    
        # 2. PRIORITAS UTAMA: PJOK JAM PERTAMA PAGI
        df_pjok = edited_df[edited_df['Kode Mapel'] == 'PJOK']
        hari_index = 0
        for _, row in df_pjok.iterrows():
            g_pjok = row['Kode Guru']
            for kelas in row['Mengajar Kelas']:
                placed = False
                attempts = 0
                while not placed and attempts < 15:
                    h_target = list_hari[hari_index % len(list_hari)]
                    hari_index += 1
                    attempts += 1
                    j_start = 2 if h_target == 'Senin' else 1
                    
                    bentrok = False
                    for step in range(3):
                        chk_j = j_start + step
                        bg = any(d['hari'] == h_target and d['jam'] == chk_j and d['kode_guru'] == g_pjok for d in plot_res)
                        bk = any(d['hari'] == h_target and d['jam'] == chk_j and d['kelas'] == kelas for d in plot_res)
                        if bg or bk: bentrok = True; break
                    if not bentrok:
                        for step in range(3):
                            plot_res.append({
                                "kode_guru": g_pjok, "kode_mapel": "PJOK",
                                "kelas": kelas, "hari": h_target, "jam": j_start + step
                            })
                        placed = True

        # 3. CRITICAL: MATEMATIKA & IPA WAJIB DI JAM AWAL (MAKSIMAL JAM KE-5)
        df_critical = edited_df[edited_df['Kode Mapel'].isin(['MAT', 'IPA'])]
        for _, row in df_critical.iterrows():
            g_name = row['Kode Guru']
            m_code = row['Kode Mapel']
            total_jp = int(row['JP/Minggu'])
            sesi = [3, 2] if total_jp == 5 else [total_jp]
            
            for kelas in row['Mengajar Kelas']:
                hari_terpakai_kelas = []
                for jp_sesi in sesi:
                    placed_sesi = False
                    for hari in list_hari:
                        if hari in hari_terpakai_kelas: continue
                        for j_mulai in range(1, 5):
                            if hari == 'Senin' and j_mulai == 1: continue
                            
                            bentrok = False
                            for step in range(jp_sesi):
                                c_j = j_mulai + step
                                bg = any(d['hari'] == hari and d['jam'] == c_j and d['kode_guru'] == g_name for d in plot_res)
                                bk = any(d['hari'] == hari and d['jam'] == c_j and d['kelas'] == kelas for d in plot_res)
                                if bg or bk: bentrok = True; break
                            if not bentrok:
                                for step in range(jp_sesi):
                                    plot_res.append({
                                        "kode_guru": g_name, "kode_mapel": m_code,
                                        "kelas": kelas, "hari": hari, "jam": j_mulai + step
                                    })
                                hari_terpakai_kelas.append(hari)
                                placed_sesi = True
                                break
                        if placed_sesi: break

        # 4. MAPEL LAINNYA SISA JAM
        df_umum = edited_df[(~edited_df['Mata Pelajaran'].str.contains("P.A.", case=False, na=False)) & (~edited_df['Kode Mapel'].isin(['PJOK', 'MAT', 'IPA', 'BK']))]
        for _, row in df_umum.iterrows():
            g_name = row['Kode Guru']
            m_code = row['Kode Mapel']
            total_jp = int(row['JP/Minggu']) if pd.notna(row['JP/Minggu']) else 0
            if total_jp <= 0 or not isinstance(row['Mengajar Kelas'], list): continue
            
            sesi = [3, 2] if total_jp == 5 else ([3, 3] if total_jp == 6 else ([2, 2] if total_jp == 4 else [total_jp]))
            for kelas in row['Mengajar Kelas']:
                hari_terpakai_kelas = []
                for jp_sesi in sesi:
                    placed_sesi = False
                    for hari in list_hari:
                        if hari in hari_terpakai_kelas: continue
                        max_jam = 6 if hari == 'Jumat' else 9
                        for j_mulai in range(1, max_jam - jp_sesi + 2):
                            if hari == 'Senin' and j_mulai == 1: continue
                            
                            bentrok = False
                            for step in range(jp_sesi):
                                c_j = j_mulai + step
                                bg = any(d['hari'] == hari and d['jam'] == c_j and d['kode_guru'] == g_name for d in plot_res)
                                bk = any(d['hari'] == hari and d['jam'] == c_j and d['kelas'] == kelas for d in plot_res)
                                if bg or bk: bentrok = True; break
                            if not bentrok:
                                for step in range(jp_sesi):
                                    plot_res.append({
                                        "kode_guru": g_name, "kode_mapel": m_code,
                                        "kelas": kelas, "hari": hari, "jam": j_mulai + step
                                    })
                                hari_terpakai_kelas.append(hari)
                                placed_sesi = True
                                break
                        if placed_sesi: break

        st.session_state.jadwal_terplot = plot_res
        st.success("✔️ Re-Optimasi Berhasil Menggunakan Data Baru!")
        st.rerun()

with tab2:
    if not st.session_state.jadwal_terplot:
        st.warning("⚠️ Silakan klik tombol 'Jalankan Engine AI Schedule Optimizer' di Tab 1 terlebih dahulu.")
    else:
        pilihan_kelas = st.selectbox("Pilih Ruang Kelas / Rombel:", semua_kelas, key="sb_kls")
        tabel_tampil = pd.DataFrame(index=list_hari, columns=[f"Jam {i}" for i in range(1, 10)]).fillna("Kosong")
        
        for h in list_hari:
            bts = 6 if h == 'Jumat' else 9
            for j in range(1, 10):
                if j > bts:
                    tabel_tampil.at[h, f"Jam {j}"] = "-"
                    continue
                if h == 'Senin' and j == 1:
                    tabel_tampil.at[h, f"Jam {j}"] = "🎗️ UPACARA"
                    continue
                    
                matches = [d for d in st.session_state.jadwal_terplot if d['hari'] == h and d['jam'] == j and d['kelas'] == pilihan_kelas]
                if matches:
                    g_singkat = ambil_tiga_digit(matches[0]['kode_guru'])
                    tabel_tampil.at[h, f"Jam {j}"] = f"{matches[0]['kode_mapel']} ({g_singkat})"
                        
        st.markdown(f"##### 📅 Preview Jadwal Teroptimasi **Kelas {pilihan_kelas}**")
        st.dataframe(tabel_tampil, use_container_width=True)
        
        excel_data = export_excel_laporan(st.session_state.jadwal_terplot, semua_kelas, list_hari)
        st.download_button(label="📥 Download Master Jadwal Excel", data=excel_data, file_name="Master_Jadwal_Teroptimasi.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

with tab3:
    st.markdown("### 🔍 Laporan Validasi Pengisian Kuota Mapel Per Rombel")
    if not st.session_state.jadwal_terplot:
        st.info("Jalankan optimasi jadwal terlebih dahulu untuk melihat hasil audit.")
    else:
        audit_records = []
        for kls in semua_kelas:
            matches_kls = [d for d in st.session_state.jadwal_terplot if d['kelas'] == kls]
            df_kls = pd.DataFrame(matches_kls)
            
            unique_mapel = df_kls['kode_mapel'].nunique() if not df_kls.empty else 0
            total_jp_terisi = len(df_kls) if not df_kls.empty else 0
            
            target_mapel = 11 
            status = "✅ LENGKAP" if unique_mapel >= target_mapel else f"⚠️ KURANG ({target_mapel - unique_mapel} Mapel)"
            
            audit_records.append({
                "Kelas": kls,
                "Jumlah Mapel Terplot": unique_mapel,
                "Total JP Terisi": total_jp_terisi,
                "Status Kelengkapan": status
            })
            
        df_audit = pd.DataFrame(audit_records)
        st.dataframe(df_audit, use_container_width=True)
